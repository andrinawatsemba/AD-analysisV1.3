import streamlit as st
import pandas as pd
import plotly.express as px
from pathlib import Path
from io import BytesIO
import sys

BASE = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE))

import config
import github_sync
import utils
from pipeline import (
    preflight_check, parse_raw_file, categorize, validate, write_warehouse,
    clear_logs, warnings_log, corrections_log, flags_log, not_aired_log,
    style_warehouse_workbook, CATEGORY_MAP, WAREHOUSE_PATH,
)

WAREHOUSE_FILE = WAREHOUSE_PATH / "warehouse.xlsx"

# ── PULL LATEST FROM GITHUB BEFORE ANYTHING ELSE ────────────────
# A fresh Streamlit Cloud container has nothing locally. This
# hydrates the local filesystem from the repo before we check
# whether the warehouse/category map exist.
if "pulled_from_github" not in st.session_state:
    github_sync.pull_file_from_github("warehouse/warehouse.xlsx", WAREHOUSE_FILE)
    github_sync.pull_file_from_github("category_map.xlsx", CATEGORY_MAP)
    st.session_state.pulled_from_github = True

# ── PAGE CONFIG ──────────────────────────────────────────────────
st.set_page_config(page_title="NBS Ad Tracker", page_icon="📺",
                    layout="wide", initial_sidebar_state="expanded")

T = config.THEME

# ── CSS: white / blue theme, red confined to the logo only ─────
# Rebuilt for readability: bigger base font, a friendlier font
# family, more breathing room in boxes/metrics. Streamlit's defaults
# run small and dense once a page has this many widgets on it.
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"] {{
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }}

    /* Base body/markdown text - was uncomfortably small by default */
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li {{
        font-size: 1.05rem !important;
        line-height: 1.6 !important;
    }}

    /* Metric cards (KPIs) - values were small relative to how much
       they matter on this page */
    [data-testid="stMetricValue"] {{ font-size: 2rem !important; font-weight: 700 !important; }}
    [data-testid="stMetricLabel"] {{ font-size: 0.95rem !important; opacity: 0.85; }}

    /* Selectbox / radio / multiselect labels and options */
    .stSelectbox label, .stRadio label, .stMultiSelect label,
    .stTextInput label {{ font-size: 1rem !important; font-weight: 600 !important; }}
    .stRadio div[role="radiogroup"] label p {{ font-size: 1rem !important; }}

    /* Dataframes/tables - default is quite small on a dense sheet */
    [data-testid="stDataFrame"] {{ font-size: 0.95rem !important; }}

    .main-header {{
        background-color: {T['PRIMARY_BLUE']};
        padding: 1.3rem 1.8rem; border-radius: 8px; margin-bottom: 1.2rem;
    }}
    .main-header h1 {{ color: white; margin: 0; font-size: 2rem; font-weight: 700; }}
    .main-header p  {{ color: rgba(255,255,255,0.9); margin: 0.2rem 0 0 0; font-size: 1rem; }}

    .section-header {{
        background-color: {T['ACCENT_BLUE']}; color: {T['DARK_GREY']};
        padding: 0.6rem 1.2rem; border-radius: 6px; font-weight: 700;
        font-size: 1.1rem; margin: 1.3rem 0 0.7rem 0;
    }}

    .gate-box, .ok-box, .warning-box, .not-aired-box {{
        padding: 1rem 1.3rem; border-radius: 6px; margin: 0.6rem 0;
        font-size: 1.05rem; line-height: 1.5;
    }}
    .gate-box    {{ background: #FFF3D6; border-left: 5px solid {T['PRIMARY_BLUE']}; }}
    .ok-box      {{ background: #E3F6E5; border-left: 5px solid #2E7D32; }}
    .warning-box {{ background: #FFF3D6; border-left: 5px solid #C9A227; }}
    .not-aired-box {{ background: {T['LIGHT_GREY']}; border-left: 5px solid {T['ACCENT_BLUE']}; }}

    div[data-testid="stSidebar"] {{ background-color: {T['PRIMARY_BLUE']}; }}
    div[data-testid="stSidebar"] * {{ color: white !important; }}
    div[data-testid="stSidebar"] label p {{ font-size: 1.05rem !important; }}
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="main-header">
    <h1>NBS Ad Tracker</h1>
    <p>Next Media Services — Internal Use Only</p>
</div>
""", unsafe_allow_html=True)

# ── HELPERS ──────────────────────────────────────────────────────

def largest_remainder_percentages(values):
    return utils.largest_remainder_percentages(values)


def count_unknowns(df):
    if df is None or df.empty or "Aired Category" not in df.columns:
        return 0
    return int((df["Aired Category"] == "****").sum())


def category_breakdown(df, ad_sq_filter=None):
    """Category x (airtime, count) table for AD-only / SQ-only /
    Combined views. Always excludes **** from the shown categories -
    per the decision that uncategorized data never appears as its
    own row anywhere; the gate ensures it's not present by the time
    a real report exists, and the Dashboard communicates
    incompleteness via the gate banner instead of a chart slice."""
    scoped = df if ad_sq_filter is None else df[df["AD/SQ"] == ad_sq_filter]
    known  = scoped[scoped["Aired Category"].isin(config.STANDARD_CATEGORIES)]

    rows = []
    for cat in config.STANDARD_CATEGORIES:
        c = known[known["Aired Category"] == cat]
        rows.append({
            "Category": cat,
            "ADs": len(c[c["AD/SQ"] == "AD"]),
            "SQs": len(c[c["AD/SQ"] == "SQ"]),
            "Airtime (secs)": int(c["Seconds Aired"].sum()),
            "Count": len(c),
        })

    airtimes = [r["Airtime (secs)"] for r in rows]
    counts   = [r["Count"] for r in rows]
    airtime_pct = largest_remainder_percentages(airtimes)
    count_pct   = largest_remainder_percentages(counts)

    for r, apct, cpct in zip(rows, airtime_pct, count_pct):
        r["% of Airtime"]    = f"{apct}%"
        r["% of Count"]      = f"{cpct}%"

    return pd.DataFrame(rows)


def sync_to_github():
    with st.spinner("Saving to persistent storage..."):
        github_sync.push_file_to_github(WAREHOUSE_FILE, "warehouse/warehouse.xlsx")
        if CATEGORY_MAP.exists():
            github_sync.push_file_to_github(CATEGORY_MAP, "category_map.xlsx")


def load_warehouse_df():
    if not WAREHOUSE_FILE.exists():
        return None
    df = pd.read_excel(WAREHOUSE_FILE, sheet_name="Clean Data")
    df.columns = df.columns.str.strip()
    return df


# ── SIDEBAR ──────────────────────────────────────────────────────
st.sidebar.markdown("### Navigation")
page = st.sidebar.radio("Navigate", [
    "Upload & Process", "Assign Categories", "Dashboard",
    "Data Table", "Not Aired Log", "Mapping Manager", "Download Reports"
], label_visibility="collapsed")

st.sidebar.markdown("---")
st.sidebar.markdown("### Warehouse Status")
_wh_df = load_warehouse_df()
if _wh_df is not None:
    st.sidebar.markdown(f"Weeks stored: **{_wh_df['Week'].nunique() if 'Week' in _wh_df.columns else 0}**")
    st.sidebar.markdown(f"Total rows: **{len(_wh_df):,}**")
    _unknowns_total = count_unknowns(_wh_df)
    if _unknowns_total > 0:
        st.sidebar.markdown(f"⚠️ **{_unknowns_total}** need categories")
else:
    st.sidebar.markdown("No warehouse found yet.")

st.sidebar.markdown("---")
st.sidebar.caption("NBS Ad Tracker v2.0")

# ── SESSION STATE ────────────────────────────────────────────────
for key, default in {"df_processed": None, "week_label": None, "pipeline_run": False}.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ════════════════════════════════════════════════════════════════
# PAGE 1 — UPLOAD & PROCESS
# ════════════════════════════════════════════════════════════════
if page == "Upload & Process":

    st.markdown('<div class="section-header">UPLOAD FILE</div>', unsafe_allow_html=True)
    st.markdown(
        "Upload either a full week (7 days side by side) or a single day's file - "
        "the pipeline detects which one you've given it automatically. "
        "Daily uploads combine into the same weekly report as you go."
    )

    uploaded = st.file_uploader("Select raw Excel file", type=["xlsx"])

    if uploaded is not None:
        temp_path = BASE / "raw_inputs" / uploaded.name
        temp_path.parent.mkdir(exist_ok=True)
        with open(temp_path, "wb") as f:
            f.write(uploaded.getbuffer())
        st.markdown(f"File received: **{uploaded.name}**")

        if st.button("Run Pipeline", type="primary"):
            clear_logs()

            with st.spinner("Running pre-flight check..."):
                result = preflight_check(temp_path)
            if not result:
                st.error("Pre-flight check failed. Check file structure.")
                st.stop()
            df_raw, days_found = result

            st.markdown('<div class="section-header">PRE-FLIGHT RESULTS</div>', unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            c1.metric("Days Detected", len(days_found))
            c2.metric("Raw Rows", df_raw.shape[0])
            c3.metric("Raw Columns", df_raw.shape[1])
            if len(days_found) == 1:
                st.info(f"Detected a single-day file ({list(days_found.keys())[0]}). "
                        "It will merge into the correct week automatically.")

            for w in warnings_log:
                st.markdown(f'<div class="warning-box">⚠ {w}</div>', unsafe_allow_html=True)

            with st.spinner("Parsing raw data..."):
                df_parsed = parse_raw_file(df_raw, days_found)
            if df_parsed is None:
                st.error("Parser failed. Check file structure.")
                st.stop()

            st.markdown('<div class="section-header">PARSER RESULTS</div>', unsafe_allow_html=True)
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Rows Extracted", f"{len(df_parsed):,}")
            c2.metric("ADs", f"{len(df_parsed[df_parsed['AD/SQ']=='AD']):,}")
            c3.metric("Squeeze Backs", f"{len(df_parsed[df_parsed['AD/SQ']=='SQ']):,}")
            c4.metric("Not Aired", f"{len(not_aired_log):,}",
                      help="Blank time or explicitly marked 'not aired' - correctly excluded, see Not Aired Log page")

            with st.spinner("Categorizing ads..."):
                df_categorized, fuzzy_matches = categorize(df_parsed, CATEGORY_MAP)
            unknown = count_unknowns(df_categorized)

            st.markdown('<div class="section-header">CATEGORIZER RESULTS</div>', unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            c1.metric("Exact Matches", len(df_categorized) - len(fuzzy_matches) - unknown)
            c2.metric("Fuzzy Matches", len(fuzzy_matches))
            c3.metric("Unknown (needs assignment)", unknown)

            with st.spinner("Validating data..."):
                df_validated = validate(df_categorized)

            with st.spinner("Writing to warehouse..."):
                success = write_warehouse(df_validated.copy(), corrections_log, flags_log, not_aired_log)

            if success:
                sync_to_github()
                st.session_state.df_processed = df_validated
                st.session_state.pipeline_run = True
                st.success(f"Pipeline complete. {len(df_validated):,} rows processed and saved.")

                if unknown > 0:
                    st.markdown(
                        f'<div class="gate-box">'
                        f'<strong>{unknown} ads still need a category</strong> before this data\'s '
                        f'report can be generated. <br>Go to <strong>Assign Categories</strong> to finish - '
                        f'it only takes a moment.</div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown('<div class="ok-box">All ads categorized - reports are ready to generate.</div>',
                                unsafe_allow_html=True)
            else:
                st.warning("Warehouse write did not complete - check the console output.")

# ════════════════════════════════════════════════════════════════
# PAGE 2 — ASSIGN CATEGORIES (the gate, made painless)
# ════════════════════════════════════════════════════════════════
elif page == "Assign Categories":

    st.markdown('<div class="section-header">ASSIGN CATEGORIES</div>', unsafe_allow_html=True)

    wh_df = load_warehouse_df()
    if wh_df is None:
        st.warning("No warehouse found. Upload and process a file first.")
        st.stop()

    unknown_mask = wh_df["Aired Category"] == "****"
    unknown_names = sorted(wh_df.loc[unknown_mask, "AD/SQ Details"].unique())

    if not unknown_names:
        st.markdown('<div class="ok-box">Nothing to assign - every ad has a category. '
                    'Reports are ready.</div>', unsafe_allow_html=True)
        st.stop()

    st.markdown(
        f'<div class="gate-box"><strong>{len(unknown_names)} unique ad name(s)</strong> '
        f'need a category ({int(unknown_mask.sum())} rows total). '
        f'Select one or more below, choose a category, and apply - reports unlock '
        f'automatically once this list is empty.</div>',
        unsafe_allow_html=True
    )

    selected = st.multiselect("Select ads to assign (pick several to batch-assign one category to all of them)",
                               unknown_names)

    chosen_category = st.selectbox("Category to apply", ["-- Select --"] + config.STANDARD_CATEGORIES)

    if st.button("Apply to Selected", type="primary", disabled=(not selected or chosen_category == "-- Select --")):
        cat_map = pd.read_excel(CATEGORY_MAP) if CATEGORY_MAP.exists() else pd.DataFrame(
            columns=["AD Name", "Category", "AD/SQ Type", "Times Seen",
                     "Category Consistent", "Last Updated", "Updated By", "Notes"])
        cat_map.columns = cat_map.columns.str.strip()

        for ad in selected:
            existing_idx = cat_map[cat_map["AD Name"].str.upper() == ad.upper()].index
            if len(existing_idx) > 0:
                cat_map.at[existing_idx[0], "Category"] = chosen_category
                cat_map.at[existing_idx[0], "Last Updated"] = pd.Timestamp.today().strftime("%d.%m.%Y")
                cat_map.at[existing_idx[0], "Updated By"] = "App User"
            else:
                ad_sq_type = wh_df.loc[wh_df["AD/SQ Details"] == ad, "AD/SQ"].iloc[0]
                new_row = pd.DataFrame([{
                    "AD Name": ad, "Category": chosen_category, "AD/SQ Type": ad_sq_type,
                    "Times Seen": 1, "Category Consistent": "Yes",
                    "Last Updated": pd.Timestamp.today().strftime("%d.%m.%Y"),
                    "Updated By": "App User", "Notes": "Added via Assign Categories"
                }])
                cat_map = pd.concat([cat_map, new_row], ignore_index=True)

            wh_df.loc[(wh_df["AD/SQ Details"] == ad) & (wh_df["Aired Category"] == "****"),
                      "Aired Category"] = chosen_category

        cat_map.to_excel(CATEGORY_MAP, index=False)

        _corr = pd.read_excel(WAREHOUSE_FILE, sheet_name="Corrections Log")
        _flags = pd.read_excel(WAREHOUSE_FILE, sheet_name="Flags Log")
        _not_aired = pd.read_excel(WAREHOUSE_FILE, sheet_name="Not Aired Log")
        with pd.ExcelWriter(WAREHOUSE_FILE, engine="openpyxl") as writer:
            wh_df.to_excel(writer, index=False, sheet_name="Clean Data")
            _corr.to_excel(writer, index=False, sheet_name="Corrections Log")
            _flags.to_excel(writer, index=False, sheet_name="Flags Log")
            _not_aired.to_excel(writer, index=False, sheet_name="Not Aired Log")
            style_warehouse_workbook(writer.book)

        sync_to_github()
        remaining = count_unknowns(wh_df)
        st.success(f"Assigned {chosen_category} to {len(selected)} ad(s). {remaining} remaining.")
        st.rerun()

# ════════════════════════════════════════════════════════════════
# PAGE 3 — DASHBOARD
# ════════════════════════════════════════════════════════════════
elif page == "Dashboard":

    wh_df = load_warehouse_df()
    if wh_df is None:
        st.warning("No warehouse found. Upload and process a file first.")
        st.stop()

    weeks_available = wh_df["Week"].unique().tolist() if "Week" in wh_df.columns else ["All Data"]
    selected_week = st.selectbox("Select week", weeks_available, index=len(weeks_available) - 1)
    week_df = wh_df[wh_df["Week"] == selected_week].copy() if "Week" in wh_df.columns else wh_df.copy()

    days_in_week = [d for d in config.DAY_ORDER if d in week_df["Date"].values]
    day_choice = st.selectbox("Select day (or view the whole week)", ["Whole Week"] + days_in_week)
    df = week_df if day_choice == "Whole Week" else week_df[week_df["Date"] == day_choice]

    unknowns = count_unknowns(df)
    if unknowns > 0:
        st.markdown(
            f'<div class="gate-box"><strong>{unknowns} ads in this view still need a category.</strong> '
            f'Numbers below reflect only what\'s been categorized so far - '
            f'go to <strong>Assign Categories</strong> to complete it.</div>',
            unsafe_allow_html=True
        )

    st.markdown('<div class="section-header">EXECUTIVE SUMMARY</div>', unsafe_allow_html=True)
    total_secs = int(df["Seconds Aired"].sum())
    c1, c2, c3 = st.columns(3)
    c1.metric("Total ADs", f"{len(df[df['AD/SQ']=='AD']):,}")
    c2.metric("Total SQs", f"{len(df[df['AD/SQ']=='SQ']):,}")
    c3.metric("Total Airtime (secs)", f"{total_secs:,}")

    st.markdown('<div class="section-header">CATEGORY BREAKDOWN — AIRTIME</div>', unsafe_allow_html=True)
    metric_choice = st.radio("Show breakdown by", ["Airtime", "Count"], horizontal=True, key="metric_toggle")
    value_col = "Airtime (secs)" if metric_choice == "Airtime" else "Count"
    pct_col   = "% of Airtime" if metric_choice == "Airtime" else "% of Count"

    view_choice = st.radio("View", ["AD only", "SQ only", "Combined (AD + SQ)"], horizontal=True)
    scope = {"AD only": "AD", "SQ only": "SQ", "Combined (AD + SQ)": None}[view_choice]
    breakdown = category_breakdown(df, ad_sq_filter=scope)

    col_left, col_right = st.columns([1, 1])
    with col_left:
        nonzero = breakdown[breakdown[value_col] > 0]
        if not nonzero.empty:
            # Labels pushed OUTSIDE the pie with leader lines, instead
            # of crammed inside thin slices where text used to overlap
            # and become unreadable (e.g. small Trade Marketing/
            # Franchise slices). Legend on the side backs it up too.
            fig = px.pie(nonzero, names="Category", values=value_col,
                         color="Category", color_discrete_map=config.CATEGORY_COLOURS,
                         title=f"{metric_choice} by Category — {view_choice}")
            fig.update_traces(
                textposition="outside", textinfo="percent",
                textfont_size=16,
                marker=dict(line=dict(color="white", width=2)),
                pull=[0.02] * len(nonzero),
            )
            fig.update_layout(
                height=480,
                margin=dict(t=60, b=40, l=40, r=40),
                legend=dict(font=dict(size=15), orientation="v", y=0.5),
                font=dict(size=15),
                title_font_size=17,
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No categorized data yet for this view.")
    with col_right:
        st.dataframe(breakdown[["Category", "ADs", "SQs", value_col, pct_col]],
                     use_container_width=True, hide_index=True)

    st.markdown('<div class="section-header">AIRTIME BY TIME BLOCK (HOURLY)</div>', unsafe_allow_html=True)
    if "Time Block" in df.columns:
        block_order = [f"{h if h else 12}{'am' if h < 12 else 'pm'}"
                       for h in range(24) for h in [h % 12 or 12]]
        # Build a clean chronological ordering of blocks actually present
        hour_seq = []
        for h in range(24):
            period = "am" if h < 12 else "pm"
            h12 = h % 12 or 12
            hour_seq.append(f"{h12}{period}")
        block_df = (df.groupby("Time Block")["Seconds Aired"].sum()
                    .reindex(hour_seq).dropna().reset_index())
        block_df.columns = ["Time Block", "Seconds Aired"]
        fig2 = px.bar(block_df, x="Time Block", y="Seconds Aired",
                      title="Airtime per Hour Block", color_discrete_sequence=[config.THEME["PRIMARY_BLUE"]])
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.info("Time Block data not available for this selection.")

# ════════════════════════════════════════════════════════════════
# PAGE 4 — DATA TABLE
# ════════════════════════════════════════════════════════════════
elif page == "Data Table":

    wh_df = load_warehouse_df()
    if wh_df is None:
        st.warning("No warehouse found. Upload and process a file first.")
        st.stop()

    st.markdown('<div class="section-header">FULL DATA TABLE</div>', unsafe_allow_html=True)

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        weeks = ["All"] + (wh_df["Week"].unique().tolist() if "Week" in wh_df.columns else [])
        sel_week = st.selectbox("Week", weeks)
    with col2:
        sel_day = st.selectbox("Day", ["All"] + config.DAY_ORDER)
    with col3:
        sel_cat = st.selectbox("Category", ["All"] + config.STANDARD_CATEGORIES + ["****"])
    with col4:
        sel_type = st.selectbox("AD / SQ", ["All", "AD", "SQ"])
    with col5:
        blocks = ["All"] + sorted(wh_df["Time Block"].dropna().unique().tolist()) if "Time Block" in wh_df.columns else ["All"]
        sel_block = st.selectbox("Time Block", blocks)

    search = st.text_input("Search by advertiser name")
    filtered = wh_df.copy()
    if sel_week != "All" and "Week" in filtered.columns:
        filtered = filtered[filtered["Week"] == sel_week]
    if sel_day != "All":
        filtered = filtered[filtered["Date"] == sel_day]
    if sel_cat != "All":
        filtered = filtered[filtered["Aired Category"] == sel_cat]
    if sel_type != "All":
        filtered = filtered[filtered["AD/SQ"] == sel_type]
    if sel_block != "All" and "Time Block" in filtered.columns:
        filtered = filtered[filtered["Time Block"] == sel_block]
    if search:
        filtered = filtered[filtered["AD/SQ Details"].str.contains(search, case=False, na=False)]

    st.markdown(f"Showing **{len(filtered):,}** of **{len(wh_df):,}** rows (sorted chronologically within each day)")

    def highlight_unknown(row):
        if row.get("Aired Category") == "****":
            return ["background-color: #FFEBEE"] * len(row)
        return [""] * len(row)

    display_cols = [c for c in [
        "Week", "Date", "Full Date", "AD/SQ Details", "Time Aired",
        "Time Block", "Seconds Aired", "Aired Category", "AD/SQ"
    ] if c in filtered.columns]

    st.dataframe(filtered[display_cols].style.apply(highlight_unknown, axis=1),
                 use_container_width=True, height=500)

# ════════════════════════════════════════════════════════════════
# PAGE 5 — NOT AIRED LOG
# ════════════════════════════════════════════════════════════════
elif page == "Not Aired Log":

    st.markdown('<div class="section-header">NOT AIRED LOG</div>', unsafe_allow_html=True)
    st.markdown(
        "Every ad/SQ with a blank time cell or explicitly marked 'not aired' - "
        "correctly excluded from airtime, but visible here instead of silently vanishing."
    )

    if not WAREHOUSE_FILE.exists():
        st.warning("No warehouse found. Upload and process a file first.")
        st.stop()

    try:
        not_aired_df = pd.read_excel(WAREHOUSE_FILE, sheet_name="Not Aired Log")
    except Exception:
        not_aired_df = pd.DataFrame()

    if not_aired_df.empty:
        st.markdown('<div class="ok-box">No not-aired entries logged.</div>', unsafe_allow_html=True)
        st.stop()

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Not-Aired", f"{len(not_aired_df):,}")
    c2.metric("AD side", f"{len(not_aired_df[not_aired_df['AD/SQ']=='AD']):,}")
    c3.metric("SQ side", f"{len(not_aired_df[not_aired_df['AD/SQ']=='SQ']):,}")

    st.markdown("**By reason:**")
    reason_counts = not_aired_df["Reason"].value_counts().reset_index()
    reason_counts.columns = ["Reason", "Count"]
    st.dataframe(reason_counts, use_container_width=True, hide_index=True)

    st.markdown("**Full list:**")
    st.dataframe(not_aired_df, use_container_width=True, height=400)

# ════════════════════════════════════════════════════════════════
# PAGE 6 — MAPPING MANAGER (admin/housekeeping only now)
# ════════════════════════════════════════════════════════════════
elif page == "Mapping Manager":

    st.markdown('<div class="section-header">MAPPING MANAGER</div>', unsafe_allow_html=True)
    st.markdown(
        "Browse and edit existing category mappings. "
        "To assign brand-new unknown ads, use **Assign Categories** instead."
    )

    if not CATEGORY_MAP.exists():
        st.error("category_map.xlsx not found.")
        st.stop()

    cat_map = pd.read_excel(CATEGORY_MAP)
    cat_map.columns = cat_map.columns.str.strip()

    search_map = st.text_input("Search ad name in category map")
    display_map = cat_map[cat_map["AD Name"].str.contains(search_map, case=False, na=False)] if search_map else cat_map

    st.markdown(f"**Category Map — {len(display_map)} entries**")
    edited_map = st.data_editor(
        display_map, use_container_width=True, height=400,
        column_config={
            "Category": st.column_config.SelectboxColumn("Category", options=config.STANDARD_CATEGORIES, required=True),
            "Last Updated": st.column_config.TextColumn("Last Updated", disabled=True),
            "Times Seen": st.column_config.NumberColumn("Times Seen", disabled=True),
        },
        disabled=["AD Name", "AD/SQ Type", "Times Seen", "Category Consistent"],
        hide_index=True, key="map_editor"
    )

    if st.button("Save Changes to Category Map", type="primary"):
        if search_map:
            cat_map.update(edited_map)
        else:
            cat_map = edited_map.copy()
        cat_map["Last Updated"] = pd.Timestamp.today().strftime("%d.%m.%Y")
        cat_map["Updated By"] = "App User"
        cat_map.to_excel(CATEGORY_MAP, index=False)
        sync_to_github()
        st.success("Category map saved.")

# ════════════════════════════════════════════════════════════════
# PAGE 7 — DOWNLOAD REPORTS (gated)
# ════════════════════════════════════════════════════════════════
elif page == "Download Reports":

    st.markdown('<div class="section-header">DOWNLOAD REPORTS</div>', unsafe_allow_html=True)

    wh_df = load_warehouse_df()
    if wh_df is None:
        st.warning("No warehouse found. Upload and process a file first.")
        st.stop()

    weeks_available = wh_df["Week"].unique().tolist() if "Week" in wh_df.columns else ["All"]
    selected_week = st.selectbox("Select week to export", weeks_available, index=len(weeks_available) - 1)
    df_export = wh_df[wh_df["Week"] == selected_week].copy() if "Week" in wh_df.columns else wh_df.copy()

    unknowns = count_unknowns(df_export)
    if unknowns > 0:
        st.markdown(
            f'<div class="gate-box"><strong>Reports are locked for this week.</strong> '
            f'{unknowns} ads still need a category - go to <strong>Assign Categories</strong> to finish, '
            f'then come back here.</div>',
            unsafe_allow_html=True
        )
        st.stop()

    st.markdown(f'<div class="ok-box">All categories assigned - this report has no unknowns.</div>',
                unsafe_allow_html=True)
    st.markdown(f"Exporting: **{selected_week}** — {len(df_export):,} rows")

    from report import write_clean_data, write_weekly_summary, write_corrections, write_flags, write_not_aired
    from pdf_report import build_pdf

    corrections = pd.read_excel(WAREHOUSE_FILE, sheet_name="Corrections Log")
    flags = pd.read_excel(WAREHOUSE_FILE, sheet_name="Flags Log")
    not_aired = pd.read_excel(WAREHOUSE_FILE, sheet_name="Not Aired Log")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### Excel Report")
        if st.button("Generate Excel Report", type="primary"):
            from openpyxl import Workbook
            output = BytesIO()
            wb = Workbook()
            wb.remove(wb.active)
            write_clean_data(wb.create_sheet("Clean Data"), df_export)
            write_weekly_summary(wb.create_sheet("Weekly Summary"), df_export, selected_week)
            write_corrections(wb.create_sheet("Corrections Log"), corrections)
            write_flags(wb.create_sheet("Flags Log"), flags)
            write_not_aired(wb.create_sheet("Not Aired Log"), not_aired)
            wb.save(output)
            output.seek(0)
            safe_week = selected_week.replace(" ", "_").replace(".", "-")
            st.download_button("Download Excel Report", data=output,
                                file_name=f"NBS_AdReport_{safe_week}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col2:
        st.markdown("### PDF Report")
        if st.button("Generate PDF Report", type="primary"):
            pdf_path = build_pdf(df_export, selected_week, corrections, flags, not_aired)
            with open(pdf_path, "rb") as f:
                pdf_bytes = f.read()
            safe_week = selected_week.replace(" ", "_").replace(".", "-")
            st.download_button("Download PDF Report", data=pdf_bytes,
                                file_name=f"NBS_AdReport_{safe_week}.pdf", mime="application/pdf")

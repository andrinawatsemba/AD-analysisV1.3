import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

import config

BASE           = Path(__file__).resolve().parent
WAREHOUSE_PATH = BASE / "warehouse" / "warehouse.xlsx"
OUTPUT_FOLDER  = BASE / "output"
OUTPUT_FOLDER.mkdir(exist_ok=True)

# ── COLORS — theme blue/white, red reserved for the logo only ──
DARK_BLUE  = config.THEME["PRIMARY_BLUE"].lstrip("#")
MID_BLUE   = config.THEME["ACCENT_BLUE"].lstrip("#")
LIGHT_BLUE = "D6F0FC"
DARK_GREY  = "404040"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"

# Chronological hour sequence for Time Block sections - NOT
# alphabetical (alphabetical would put "10am" before "6am").
HOUR_SEQUENCE = []
for _h in range(24):
    _period = "am" if _h < 12 else "pm"
    _h12 = _h % 12 or 12
    HOUR_SEQUENCE.append(f"{_h12}{_period}")

# ── STYLE HELPERS ────────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, bg=DARK_BLUE, fg=WHITE, size=11):
    cell.fill = make_fill(bg)
    cell.font = Font(bold=True, color=fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border()

def style_cell(cell, bold=False, align="center", bg=None, fg=DARK_GREY):
    if bg:
        cell.fill = make_fill(bg)
    cell.font = Font(bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = make_border()

def auto_width(ws, max_width=45):
    for col in ws.columns:
        max_len, col_letter = 0, None
        for cell in col:
            if hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)

# ── LOAD WAREHOUSE ───────────────────────────────────────────────
def load_warehouse(week_label=None):
    if not WAREHOUSE_PATH.exists():
        print("[ERROR] Warehouse not found. Run pipeline.py first.")
        return None, None, None, None, None, None

    df = pd.read_excel(WAREHOUSE_PATH, sheet_name="Clean Data")
    df.columns = df.columns.str.strip()
    corrections = pd.read_excel(WAREHOUSE_PATH, sheet_name="Corrections Log")
    flags       = pd.read_excel(WAREHOUSE_PATH, sheet_name="Flags Log")
    try:
        dropped_rows = pd.read_excel(WAREHOUSE_PATH, sheet_name="Dropped Rows Log")
    except Exception:
        dropped_rows = pd.DataFrame(columns=["Row Reference", "AD/SQ Details", "AD/SQ", "Reason"])
    try:
        day_detection = pd.read_excel(WAREHOUSE_PATH, sheet_name="Day Detection Log")
    except Exception:
        day_detection = pd.DataFrame(columns=["Day", "Column (as in Excel)", "Row (as in Excel)",
                                               "Date Assigned", "Full Date", "Raw Entries Scanned", "Week"])

    if week_label and "Week" in df.columns:
        df = df[df["Week"] == week_label]
    elif "Week" in df.columns:
        week_label = df["Week"].iloc[-1]
        df = df[df["Week"] == week_label]

    if "Week" in day_detection.columns:
        day_detection = day_detection[day_detection["Week"] == week_label]

    return df, week_label, corrections, flags, dropped_rows, day_detection

# ── SHEET: CLEAN DATA ────────────────────────────────────────────
# Columns: A=Date, B=Full Date, C=AD/SQ Details, D=Time Aired,
# E=Time Block, F=Seconds Aired, G=Aired Category, H=AD/SQ
def write_clean_data(ws, df):
    headers = ["Date", "Full Date", "AD/SQ Details", "Time Aired",
               "Time Block", "Seconds Aired", "Aired Category", "AD/SQ"]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        bg = LIGHT_GREY if row_idx % 2 == 0 else WHITE
        values = [row.get(h, "") for h in headers]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            style_cell(cell, bg=bg, align="left" if col == 3 else "center")

    # Category dropdown on column G - no **** option here: by the
    # time a report is generated, the gate guarantees every row
    # already has a real category assigned.
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(config.STANDARD_CATEGORIES) + '"',
        allow_blank=False, showDropDown=False
    )
    dv.sqref = f"G2:G{len(df) + 1}"
    ws.add_data_validation(dv)

    auto_width(ws)
    ws.freeze_panes = "A2"
    print(f"   [OK] Clean Data sheet written - {len(df)} rows")

# ── SHEET: WEEKLY SUMMARY ────────────────────────────────────────
def write_weekly_summary(ws, df, week_label):
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "NBS AD TRACKING - WEEKLY SUMMARY"
    title.font = Font(bold=True, size=14, color=DARK_BLUE)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"Week: {week_label}    |    Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sub.font = Font(italic=True, color=DARK_GREY, size=10)
    ws.row_dimensions[2].height = 18

    row_cursor = 4

    # ── Section 1: Ads aired per day ──
    ws.merge_cells(f"A{row_cursor}:H{row_cursor}")
    sec1 = ws[f"A{row_cursor}"]
    sec1.value = "SECTION 1 - ADS AIRED PER DAY"
    style_header(sec1, bg=DARK_BLUE, size=11)
    ws.row_dimensions[row_cursor].height = 20
    row_cursor += 1

    s1_headers = ["Day", "Total ADs", "Total SQs", "Total Entries", "Airtime (secs)", "", "", ""]
    for col, h in enumerate(s1_headers, start=1):
        style_header(ws.cell(row=row_cursor, column=col, value=h), bg=MID_BLUE)
    row_cursor += 1

    days = [d for d in config.DAY_ORDER if d in df["Date"].values]
    section1_start = row_cursor
    for i, day in enumerate(days):
        row = row_cursor
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws.cell(row=row, column=1, value=day)
        style_cell(ws.cell(row=row, column=1), bold=True, align="left", bg=bg)
        ws.cell(row=row, column=2, value=f"=COUNTIFS('Clean Data'!A:A,\"{day}\",'Clean Data'!H:H,\"AD\")")
        ws.cell(row=row, column=3, value=f"=COUNTIFS('Clean Data'!A:A,\"{day}\",'Clean Data'!H:H,\"SQ\")")
        ws.cell(row=row, column=4, value=f"=COUNTIF('Clean Data'!A:A,\"{day}\")")
        ws.cell(row=row, column=5, value=f"=SUMIF('Clean Data'!A:A,\"{day}\",'Clean Data'!F:F)")
        for col in range(2, 6):
            style_cell(ws.cell(row=row, column=col), bg=bg)
        row_cursor += 1

    total_row = row_cursor
    ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    style_cell(ws.cell(row=total_row, column=1), bold=True, bg=DARK_BLUE, fg=WHITE)
    for col, letter in zip(range(2, 6), ["B", "C", "D", "E"]):
        cell = ws.cell(row=total_row, column=col,
                        value=f"=SUM({letter}{section1_start}:{letter}{total_row - 1})")
        style_cell(cell, bold=True, bg=DARK_BLUE, fg=WHITE)
    row_cursor = total_row + 3

    # ── Section 2: Category breakdown - AD / SQ / Combined ──
    # Three sub-tables per your "show both, clearly" decision - not
    # merged into one ambiguous table.
    def write_category_section(row_cursor, title, ad_sq_filter):
        ws.merge_cells(f"A{row_cursor}:H{row_cursor}")
        sec = ws[f"A{row_cursor}"]
        sec.value = title
        style_header(sec, bg=DARK_BLUE, size=11)
        ws.row_dimensions[row_cursor].height = 20
        row_cursor += 1

        headers = ["Category", "Airtime (secs)", "% of Airtime", "", "", "", "", ""]
        for col, h in enumerate(headers, start=1):
            style_header(ws.cell(row=row_cursor, column=col, value=h), bg=MID_BLUE)
        row_cursor += 1

        section_start = row_cursor
        filter_clause = f",'Clean Data'!H:H,\"{ad_sq_filter}\"" if ad_sq_filter else ""
        for i, cat in enumerate(config.STANDARD_CATEGORIES):
            row = row_cursor
            bg = LIGHT_BLUE if i % 2 == 0 else WHITE
            ws.cell(row=row, column=1, value=cat)
            style_cell(ws.cell(row=row, column=1), bold=True, align="left", bg=bg)
            ws.cell(row=row, column=2,
                    value=f"=SUMIFS('Clean Data'!F:F,'Clean Data'!G:G,\"{cat}\"{filter_clause})")
            end_row = section_start + len(config.STANDARD_CATEGORIES) - 1
            ws.cell(row=row, column=3, value=f"=B{row}/SUM(B{section_start}:B{end_row})")
            ws.cell(row=row, column=3).number_format = "0.0%"
            for col in range(2, 4):
                style_cell(ws.cell(row=row, column=col), bg=bg)
            row_cursor += 1

        cat_total_row = row_cursor
        ws.cell(row=cat_total_row, column=1, value="TOTAL")
        style_cell(ws.cell(row=cat_total_row, column=1), bold=True, bg=DARK_BLUE, fg=WHITE)
        cell = ws.cell(row=cat_total_row, column=2,
                        value=f"=SUM(B{section_start}:B{cat_total_row - 1})")
        style_cell(cell, bold=True, bg=DARK_BLUE, fg=WHITE)
        ws.cell(row=cat_total_row, column=3, value="100.0%")
        style_cell(ws.cell(row=cat_total_row, column=3), bold=True, bg=DARK_BLUE, fg=WHITE)
        return cat_total_row + 3

    row_cursor = write_category_section(row_cursor, "SECTION 2A - CATEGORY BREAKDOWN (ADs ONLY)", "AD")
    row_cursor = write_category_section(row_cursor, "SECTION 2B - CATEGORY BREAKDOWN (SQs ONLY)", "SQ")
    row_cursor = write_category_section(row_cursor, "SECTION 2C - CATEGORY BREAKDOWN (COMBINED)", None)

    # ── Section 3: Airtime by Time Block (chronological, not A-Z) ──
    ws.merge_cells(f"A{row_cursor}:H{row_cursor}")
    sec3 = ws[f"A{row_cursor}"]
    sec3.value = "SECTION 3 - AIRTIME BY TIME BLOCK"
    style_header(sec3, bg=DARK_BLUE, size=11)
    ws.row_dimensions[row_cursor].height = 20
    row_cursor += 1

    headers3 = ["Time Block", "Airtime (secs)", "", "", "", "", "", ""]
    for col, h in enumerate(headers3, start=1):
        style_header(ws.cell(row=row_cursor, column=col, value=h), bg=MID_BLUE)
    row_cursor += 1

    blocks_present = [b for b in HOUR_SEQUENCE if b in df["Time Block"].values] if "Time Block" in df.columns else []
    for i, block in enumerate(blocks_present):
        row = row_cursor
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        ws.cell(row=row, column=1, value=block)
        style_cell(ws.cell(row=row, column=1), bold=True, align="left", bg=bg)
        ws.cell(row=row, column=2, value=f"=SUMIF('Clean Data'!E:E,\"{block}\",'Clean Data'!F:F)")
        style_cell(ws.cell(row=row, column=2), bg=bg)
        row_cursor += 1

    auto_width(ws)
    ws.freeze_panes = "A3"
    print("   [OK] Weekly Summary sheet written")

# ── SHEET: CORRECTIONS / FLAGS / NOT AIRED LOGS ─────────────────
def _write_log_sheet(ws, data, headers, empty_message):
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    if data is None or len(data) == 0:
        ws.cell(row=2, column=1, value=empty_message)
        style_cell(ws.cell(row=2, column=1), align="left")
    else:
        for row_idx, (_, row) in enumerate(data.iterrows(), start=2):
            bg = LIGHT_GREY if row_idx % 2 == 0 else WHITE
            for col, field in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=row.get(field, ""))
                style_cell(cell, bg=bg, align="left")
    auto_width(ws)

def write_corrections(ws, corrections):
    _write_log_sheet(ws, corrections,
                      ["Row Reference", "Original Value", "Corrected To", "Reason"],
                      "No corrections were made this week")
    print(f"   [OK] Corrections Log sheet written - {len(corrections) if corrections is not None else 0} entries")

def write_flags(ws, flags):
    _write_log_sheet(ws, flags,
                      ["Row Reference", "AD/SQ Details", "Issue", "Suggestion"],
                      "No flags were raised this week")
    print(f"   [OK] Flags Log sheet written - {len(flags) if flags is not None else 0} entries")

def write_dropped_rows(ws, dropped_rows):
    _write_log_sheet(ws, dropped_rows,
                      ["Row Reference", "AD/SQ Details", "AD/SQ", "Reason", "Day"],
                      "Nothing was dropped this week - every named ad/SQ made it into Clean Data")
    print(f"   [OK] Dropped Rows Log sheet written - {len(dropped_rows) if dropped_rows is not None else 0} entries")

# ── SHEET: VALIDATION ────────────────────────────────────────────
# What this answers, in plain terms: "did the cleaning step throw
# away anything it shouldn't have?" Not the arithmetic (that's the
# Weekly Summary's job, via visible native formulas) - this is
# specifically about whether every row in the raw file is accounted
# for. Two checks, both things you can verify against your own copy
# of the raw file without touching any code.
def write_validation_sheet(ws, df, dropped_rows, day_detection, week_label):
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "VALIDATION - DID THE CLEANING STEP DROP ANYTHING IT SHOULDN'T HAVE?"
    title.font = Font(bold=True, size=13, color=DARK_BLUE)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = (f"Week: {week_label}    |    Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    sub.font = Font(italic=True, color=DARK_GREY, size=10)

    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    sec = ws[f"A{row}"]
    sec.value = "CHECK 1 - ROW COUNT RECONCILIATION"
    style_header(sec, bg=DARK_BLUE, size=11)
    ws.row_dimensions[row].height = 20
    row += 2

    raw_scanned  = int(day_detection["Raw Entries Scanned"].sum()) if not day_detection.empty else 0
    clean_rows   = len(df)
    dropped_n    = len(dropped_rows) if dropped_rows is not None else 0

    explain = [
        "Every AD/SQ name found in the raw file lands in exactly one of two places:",
        "Clean Data (it had a valid time and wasn't operational text), or the Dropped",
        "Rows Log (blank time, marked 'not aired', or operational/schedule text - not",
        "a real ad). If the two add up to the raw total below, nothing was lost.",
    ]
    for line in explain:
        ws.cell(row=row, column=1, value=line).font = Font(size=10, color=DARK_GREY)
        row += 1
    row += 1

    labels = ["Raw AD/SQ entries found in the file", "Rows in Clean Data (this week)",
              "Rows in Dropped Rows Log (this week)", "Clean Data + Dropped Rows Log"]
    values = [raw_scanned, clean_rows, dropped_n, clean_rows + dropped_n]
    for label, val in zip(labels, values):
        ws.cell(row=row, column=1, value=label)
        style_cell(ws.cell(row=row, column=1), align="left", bold=True)
        ws.cell(row=row, column=2, value=val)
        style_cell(ws.cell(row=row, column=2), align="center")
        row += 1

    check_row = row
    ws.cell(row=check_row, column=1, value="Check: does raw total = Clean Data + Dropped Rows Log?")
    style_cell(ws.cell(row=check_row, column=1), align="left", bold=True, bg=LIGHT_BLUE)
    match = (raw_scanned == clean_rows + dropped_n)
    ws.cell(row=check_row, column=2,
            value=("YES - MATCH" if match else "NO - MISMATCH, INVESTIGATE"))
    style_cell(ws.cell(row=check_row, column=2), align="center", bold=True,
               bg=(LIGHT_BLUE if match else "FFC7CE"))
    row = check_row + 3

    # ── Check 2: Day & Date Detection ──
    ws.merge_cells(f"A{row}:F{row}")
    sec2 = ws[f"A{row}"]
    sec2.value = "CHECK 2 - DID THE TOOL FIND THE RIGHT DAY IN THE RIGHT PLACE?"
    style_header(sec2, bg=DARK_BLUE, size=11)
    ws.row_dimensions[row].height = 20
    row += 1

    ws.cell(row=row, column=1,
            value="Open your original raw file and confirm each day's label really is where this says it is.")
    ws.cell(row=row, column=1).font = Font(size=10, color=DARK_GREY, italic=True)
    row += 2

    headers = ["Day", "Found at Cell", "Date Assigned", "Raw Entries This Day"]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=row, column=col, value=h), bg=MID_BLUE)
    row += 1

    if not day_detection.empty:
        detection_sorted = day_detection.set_index("Day").reindex(
            [d for d in config.DAY_ORDER if d in day_detection["Day"].values]
        ).reset_index()
        for i, r in detection_sorted.iterrows():
            bg = LIGHT_BLUE if i % 2 == 0 else WHITE
            cell_ref = f"{r['Column (as in Excel)']}{r['Row (as in Excel)']}"
            values = [r["Day"], cell_ref, r["Date Assigned"], r["Raw Entries Scanned"]]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                style_cell(cell, bg=bg, align="left" if col <= 2 else "center")
            row += 1

    auto_width(ws)
    print("   [OK] Validation sheet written")

# ── MAIN ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n" + "="*50)
    print("REPORT BUILDER")
    print("="*50)

    result = load_warehouse()
    if result[0] is None:
        exit()
    df, week_label, corrections, flags, dropped_rows, day_detection = result

    safe_week = week_label.replace(" ", "_").replace(".", "-")
    output_file = OUTPUT_FOLDER / f"NBS_AdReport_{safe_week}.xlsx"

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Validation sheet first - the first thing anyone sees when they
    # open the file, per your request to put it "at the beginning".
    ws_validation = wb.create_sheet("Validation")
    ws_clean = wb.create_sheet("Clean Data")
    ws_summary = wb.create_sheet("Weekly Summary")
    ws_corr = wb.create_sheet("Corrections Log")
    ws_flags = wb.create_sheet("Flags Log")
    ws_dropped = wb.create_sheet("Dropped Rows Log")

    print("\nWriting sheets...")
    write_validation_sheet(ws_validation, df, dropped_rows, day_detection, week_label)
    write_clean_data(ws_clean, df)
    write_weekly_summary(ws_summary, df, week_label)
    write_corrections(ws_corr, corrections)
    write_flags(ws_flags, flags)
    write_dropped_rows(ws_dropped, dropped_rows)

    wb.save(output_file)
    print(f"\n[OK] Report saved to: {output_file}")
    print("="*50)


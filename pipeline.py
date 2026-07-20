import pandas as pd
import re
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter

import config

# ── PATHS ────────────────────────────────────────────────────────
BASE           = Path(__file__).resolve().parent
RAW_FOLDER     = BASE / "raw_inputs"
WAREHOUSE_PATH = BASE / "warehouse"
CATEGORY_MAP   = BASE / "category_map.xlsx"
OUTPUT_FOLDER  = BASE / "output"

WAREHOUSE_PATH.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

# ── LOGGING ──────────────────────────────────────────────────────
# dropped_rows_log is a COMPLETE accounting of every named AD/SQ cell
# that did NOT make it into Clean Data, whatever the reason - blank
# time, explicit "not aired", or operational/non-ad text (shift
# notes, program names). Every one of the three used to be either
# invisible (operational text) or under-counted (SQ not-aired had no
# logging at all). Now every one is logged with why, which is what
# makes the row-count reconciliation in the Validation sheet honest:
# raw entries found = Clean Data rows + dropped_rows_log rows,
# exactly, with nothing unaccounted for in between.
warnings_log      = []
corrections_log   = []
flags_log         = []
dropped_rows_log  = []

# New: exactly where each day/date was found in the raw file (column
# letter + row number, the way you'd see it in Excel), plus how many
# raw AD/SQ name cells were scanned that day - so you can manually
# verify "Tuesday really was column M, row 1" against your own copy
# of the file, and reconcile row counts (Clean Data + Dropped Rows =
# Raw Entries Scanned, exactly).
day_detection_log = []

def log_warning(message):
    warnings_log.append(message)
    print(f"[WARNING] {message}")

def log_correction(row_ref, original, corrected, reason):
    corrections_log.append({
        "Row Reference": row_ref,
        "Original Value": original,
        "Corrected To":   corrected,
        "Reason":         reason
    })

def log_flag(row_ref, detail, issue, suggestion="Review manually"):
    flags_log.append({
        "Row Reference":  row_ref,
        "AD/SQ Details":  detail,
        "Issue":          issue,
        "Suggestion":     suggestion
    })

def log_dropped_row(row_ref, detail, ad_sq, reason, day=None, full_date=None):
    dropped_rows_log.append({
        "Row Reference": row_ref,
        "AD/SQ Details": detail,
        "AD/SQ":         ad_sq,
        # "No time entered" / "Marked not aired" /
        # "Excluded - operational or schedule text, not an ad"
        "Reason":        reason,
        "Day":           day,
        "Full Date":     full_date,
    })

def clear_logs():
    warnings_log.clear()
    corrections_log.clear()
    flags_log.clear()
    dropped_rows_log.clear()
    day_detection_log.clear()

# ── NON-AD PHRASE CHECK ──────────────────────────────────────────
# Unchanged from the original - verified against your real raw file
# that substring matching is correct here (catches staff annotations
# like "END OF SHIFT SIGN:..." and "NOTE:..." that carry trailing
# text), not a bug.
def is_non_ad_phrase(value):
    if not value:
        return False
    cleaned = str(value).strip().lower()
    return any(phrase in cleaned for phrase in config.NON_AD_PHRASES)

# ── HELPER: NORMALIZE DAY NAME ───────────────────────────────────
def normalize_day(value):
    if not value or str(value).strip().lower() in ["nan", "none", ""]:
        return None
    cleaned = str(value).strip().lower()
    return config.DAY_ALIASES.get(cleaned, None)

# ── HELPER: EXTRACT DATE ─────────────────────────────────────────
def extract_date(value):
    if not value or str(value).strip().lower() in ["nan", "none", ""]:
        return None
    value = str(value).strip()
    patterns = [
        r"\d{1,2}\.\d{1,2}\.\d{4}",
        r"\d{1,2}/\d{1,2}/\d{4}",
        r"\d{4}-\d{1,2}-\d{1,2}"
    ]
    for pattern in patterns:
        match = re.search(pattern, value)
        if match:
            try:
                date = pd.to_datetime(match.group(), dayfirst=True)
                return date.strftime("%d.%m.%Y")
            except Exception:
                continue
    return None

# ── HELPER: PARSE TIME ───────────────────────────────────────────
# This replaces the old normalize_time(). It returns BOTH the display
# string (unchanged format, e.g. "06:15am") AND minutes-since-midnight.
# The minutes value is new - it's what makes chronological sorting
# and Time Block derivation actually correct, instead of the old
# alphabetical string sort which put 12:xxam (midnight) in the wrong
# position.
#
# Returns (display_string, minutes) or (None, None) if the cell has
# no usable time - covers blank cells AND the literal text
# "not aired", since your rule treats both the same way.
def parse_time(value):
    if value is None:
        return None, None
    raw = str(value).strip()
    if not raw or raw.lower() in ("nan", "none", ""):
        return None, None
    if "not aired" in raw.lower():
        return None, None

    # Excel stores some cells as a decimal fraction of a day
    # (e.g. 0.25 == 6:00am). Try that first.
    try:
        decimal = float(raw)
        if 0 <= decimal < 1:
            total_minutes = int(round(decimal * 24 * 60))
            hours   = total_minutes // 60
            minutes = total_minutes % 60
            period    = "am" if hours < 12 else "pm"
            hours_12  = hours % 12 or 12
            display = f"{hours_12:02d}:{minutes:02d}{period}"
            return display, total_minutes
    except (ValueError, TypeError):
        pass

    cleaned = raw.replace(";", ":")
    cleaned = re.sub(r"\s*(am|pm)\s*", r"\1", cleaned.lower())

    match = re.search(r"(\d{1,2}):(\d{2})(am|pm)?", cleaned)
    if not match:
        return None, None

    hour_str, minute_str, period = match.groups()
    hour   = int(hour_str)
    minute = int(minute_str)
    if period is None:
        # No am/pm marker present - keep prior lenient behavior of
        # still accepting the value, guessing period from the hour.
        period = "am" if hour < 12 else "pm"

    hour12  = hour % 12 or 12
    display = f"{hour12:02d}:{minute:02d}{period}"

    hour24 = hour % 12
    if period == "pm":
        hour24 += 12
    total_minutes = hour24 * 60 + minute

    return display, total_minutes

# ── HELPER: TIME BLOCK ───────────────────────────────────────────
# Derives the hourly bucket from minutes-since-midnight, e.g.
# 9:07am -> "9am". Floors to the start of the hour.
def time_block_from_minutes(minutes):
    if minutes is None:
        return None
    hour24  = (minutes // 60) % 24
    period  = "am" if hour24 < 12 else "pm"
    hour12  = hour24 % 12 or 12
    return f"{hour12}{period}"

# ── HELPER: CALENDAR WEEK LABEL ──────────────────────────────────
# Fixes the daily-upload bug: the old code labeled each upload's week
# as "{min date in THIS upload} to {max date in THIS upload}", which
# meant a single day's upload got its own tiny "week". This instead
# computes the Monday-to-Sunday calendar week that a date belongs to,
# independent of what else is in the current upload - so Monday's
# solo upload and Sunday's solo upload land in the SAME week and
# merge correctly in the warehouse.
def calendar_week_label(date_str):
    """date_str is 'dd.mm.yyyy' (the Full Date format used throughout)."""
    d = datetime.strptime(date_str, "%d.%m.%Y")
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return f"{monday.strftime('%d.%m.%Y')} to {sunday.strftime('%d.%m.%Y')}"

# ── PRE-FLIGHT CHECK ─────────────────────────────────────────────
def preflight_check(filepath):
    print("\n" + "="*50)
    print("PRE-FLIGHT CHECK")
    print("="*50)

    if not filepath.exists():
        print(f"[ERROR] File not found: {filepath.name}")
        return False

    if filepath.suffix.lower() != ".xlsx":
        log_warning(f"File is {filepath.suffix} not .xlsx - may cause issues")
    else:
        print("[OK] File format: .xlsx")

    try:
        df_raw = pd.read_excel(filepath, header=None)
        print(f"[OK] File loaded: {df_raw.shape[0]} rows x {df_raw.shape[1]} columns")
    except Exception as e:
        print(f"[ERROR] Could not read file: {e}")
        return False

    days_found = {}
    for row_idx in range(min(4, df_raw.shape[0])):
        for col_idx in range(df_raw.shape[1]):
            cell = df_raw.iloc[row_idx, col_idx]
            day  = normalize_day(cell)
            if day and day not in days_found:
                days_found[day] = {
                    "row":      row_idx,
                    "col":      col_idx,
                    "original": str(cell).strip()
                }

    if not days_found:
        print("[ERROR] No day names detected in first 4 rows.")
        log_warning("No day names found - file may have unexpected structure")
    else:
        # No assumption of exactly 7 days - a single-day file with one
        # day block detected here works the same way through the rest
        # of the pipeline as a full week does.
        print(f"[OK] Days detected ({len(days_found)}): {', '.join(days_found.keys())}")

    dates_found = {}
    for row_idx in range(min(4, df_raw.shape[0])):
        for col_idx in range(df_raw.shape[1]):
            cell = df_raw.iloc[row_idx, col_idx]
            date = extract_date(cell)
            if date:
                dates_found[col_idx] = date

    if dates_found:
        print(f"[OK] Dates detected: {', '.join(dates_found.values())}")
    else:
        log_warning("No dates detected in first 4 rows - dates will be missing from output")

    if CATEGORY_MAP.exists():
        cat_map = pd.read_excel(CATEGORY_MAP)
        print(f"[OK] Category map found: {len(cat_map)} ads mapped")
    else:
        log_warning("category_map.xlsx not found - all ads will be assigned ****")

    print("\n" + "-"*50)
    if warnings_log:
        print(f"[WARNING] {len(warnings_log)} warning(s) found - continuing")
    else:
        print("[OK] Pre-flight check passed with no warnings")
    print("-"*50)

    return df_raw, days_found

# ── PARSER ───────────────────────────────────────────────────────
def parse_raw_file(df_raw, days_found):
    print("\n" + "="*50)
    print("PARSER")
    print("="*50)

    all_rows    = []
    known_dates = {}

    for day, info in days_found.items():
        for r in range(max(0, info["row"] - 1), min(info["row"] + 4, df_raw.shape[0])):
            for c in range(max(0, info["col"] - 1), min(info["col"] + 6, df_raw.shape[1])):
                found = extract_date(df_raw.iloc[r, c])
                if found:
                    known_dates[day] = found
                    break
            if day in known_dates:
                break

    for day, info in days_found.items():
        day_col    = info["col"]
        header_row = info["row"]

        date_val = known_dates.get(day)

        if not date_val:
            inferred = None
            for known_day, known_date in known_dates.items():
                known_idx   = config.DAY_ORDER.index(known_day) if known_day in config.DAY_ORDER else -1
                current_idx = config.DAY_ORDER.index(day) if day in config.DAY_ORDER else -1
                if known_idx >= 0 and current_idx >= 0:
                    delta = current_idx - known_idx
                    try:
                        base     = pd.to_datetime(known_date, dayfirst=True)
                        inferred = (base + pd.Timedelta(days=delta)).strftime("%d.%m.%Y")
                        break
                    except Exception:
                        continue
            if inferred:
                date_val = inferred
                log_warning(f"Date for {day} was missing - inferred as {inferred}")
                log_correction(f"{day} header", "No date", inferred,
                               "Inferred from surrounding dates")
            else:
                log_warning(f"No date found for {day} - left blank")
                date_val = ""
        else:
            print(f"   Date: {date_val}")

        detection_entry = {
            "Day": day,
            "Column (as in Excel)": get_column_letter(day_col + 1),
            "Row (as in Excel)": header_row + 1,
            "Date Assigned": date_val or "(none)",
            "Full Date": date_val,
            "Raw Entries Scanned": 0,   # filled in after this day's rows are processed
        }
        day_detection_log.append(detection_entry)

        ad_name_col = day_col + 1
        ad_time_col = None
        sq_name_col = None
        sq_time_col = None

        for r in range(header_row, min(header_row + 4, df_raw.shape[0])):
            for c in range(day_col, min(day_col + 8, df_raw.shape[1])):
                cell       = str(df_raw.iloc[r, c]).strip().lower()
                normalized = config.HEADER_ALIASES.get(cell, cell)

                if normalized == "TIME AIRED" and ad_time_col is None and c > day_col:
                    ad_time_col = c
                elif normalized == "SQUEEZE BACKS" and sq_name_col is None:
                    sq_name_col = c
                elif normalized == "TIME AIRED" and ad_time_col is not None and sq_time_col is None:
                    sq_time_col = c

        if ad_time_col is None:
            ad_time_col = day_col + 2
            log_warning(f"{day}: could not detect AD time column by header - "
                        f"falling back to column position {ad_time_col}. Verify this day's layout.")
        if sq_name_col is None:
            sq_name_col = day_col + 3
            log_warning(f"{day}: could not detect Squeeze Back column by header - "
                        f"falling back to column position {sq_name_col}. Verify this day's layout.")
        if sq_time_col is None:
            sq_time_col = day_col + 4
            log_warning(f"{day}: could not detect SQ time column by header - "
                        f"falling back to column position {sq_time_col}. Verify this day's layout.")

        print(f"   AD name col: {ad_name_col}, AD time col: {ad_time_col}")
        print(f"   SQ name col: {sq_name_col}, SQ time col: {sq_time_col}")

        ad_count, sq_count = 0, 0
        raw_entries_this_day = 0

        for row_idx in range(header_row + 1, df_raw.shape[0]):

            # ── AD side ──
            try:
                ad_name_raw = str(df_raw.iloc[row_idx, ad_name_col]).strip()
                ad_time_raw = df_raw.iloc[row_idx, ad_time_col]
            except Exception:
                ad_name_raw = ""
                ad_time_raw = None

            ad_display, ad_minutes = parse_time(ad_time_raw)
            row_ref = f"Row {row_idx} ({day})"

            if ad_name_raw and ad_name_raw.lower() not in ["nan", "none", ""]:
                raw_entries_this_day += 1
                if is_non_ad_phrase(ad_name_raw):
                    # Previously silent - now logged, so it counts
                    # toward the reconciliation instead of vanishing.
                    log_dropped_row(row_ref, ad_name_raw.upper().strip(), "AD",
                                     "Excluded - operational or schedule text, not an ad",
                                     day=day, full_date=date_val)
                elif ad_minutes is not None:
                    all_rows.append({
                        "Date":           day,
                        "Full Date":      date_val,
                        "AD/SQ Details":  ad_name_raw.upper().strip(),
                        "Time Aired":     ad_display,
                        "_minutes":       ad_minutes,
                        "Seconds Aired":  config.AD_SECONDS,
                        "Aired Category": "****",
                        "AD/SQ":          "AD"
                    })
                    ad_count += 1
                else:
                    # No usable time - either blank cell or literal
                    # "not aired". Both mean the same thing per your
                    # rule, logged identically but with the real reason.
                    reason = ("Marked not aired"
                              if "not aired" in str(ad_time_raw).strip().lower()
                              else "No time entered")
                    log_dropped_row(row_ref, ad_name_raw.upper().strip(), "AD", reason, day=day, full_date=date_val)

            # ── SQ side ──
            try:
                sq_name_raw = str(df_raw.iloc[row_idx, sq_name_col]).strip()
                sq_time_raw = df_raw.iloc[row_idx, sq_time_col]
            except Exception:
                sq_name_raw = ""
                sq_time_raw = None

            sq_display, sq_minutes = parse_time(sq_time_raw)

            if sq_name_raw and sq_name_raw.lower() not in ["nan", "none", ""]:
                raw_entries_this_day += 1
                if is_non_ad_phrase(sq_name_raw):
                    log_dropped_row(row_ref, sq_name_raw.upper().strip(), "SQ",
                                     "Excluded - operational or schedule text, not an ad",
                                     day=day, full_date=date_val)
                elif sq_minutes is not None:
                    all_rows.append({
                        "Date":           day,
                        "Full Date":      date_val,
                        "AD/SQ Details":  sq_name_raw.upper().strip(),
                        "Time Aired":     sq_display,
                        "_minutes":       sq_minutes,
                        "Seconds Aired":  config.SQ_SECONDS,
                        "Aired Category": "****",
                        "AD/SQ":          "SQ"
                    })
                    sq_count += 1
                else:
                    # Symmetric with the AD side above - the old code
                    # had no equivalent branch here at all, so SQ
                    # not-aired/blank-time rows vanished uncounted.
                    reason = ("Marked not aired"
                              if "not aired" in str(sq_time_raw).strip().lower()
                              else "No time entered")
                    log_dropped_row(row_ref, sq_name_raw.upper().strip(), "SQ", reason, day=day, full_date=date_val)

        detection_entry["Raw Entries Scanned"] = raw_entries_this_day
        print(f"   ADs extracted: {ad_count}  |  SQs extracted: {sq_count}  |  "
              f"Raw entries scanned: {raw_entries_this_day}")

    if not all_rows:
        print("\n[ERROR] No rows extracted.")
        return None

    df = pd.DataFrame(all_rows)

    # Chronological sort using real minutes-since-midnight, not the
    # old alphabetical string sort (which misplaced 12:00am-12:59am
    # entries at the end of the morning instead of the start of the
    # day). Time Block is derived from the same minutes value.
    df["Date"] = pd.Categorical(df["Date"], categories=config.DAY_ORDER, ordered=True)
    df = df.sort_values(["Date", "_minutes"]).reset_index(drop=True)
    df["Time Block"] = df["_minutes"].apply(time_block_from_minutes)
    df = df.drop(columns=["_minutes"])

    print("\n" + "="*50)
    print("PARSER COMPLETE")
    print(f"   Total rows extracted: {len(df)}")
    print(f"   ADs: {len(df[df['AD/SQ'] == 'AD'])}  |  SQs: {len(df[df['AD/SQ'] == 'SQ'])}")
    print(f"   Dropped rows logged: {len(dropped_rows_log)}")
    print("="*50)

    return df

# ── CATEGORIZER ──────────────────────────────────────────────────
# Behavior unchanged from the original by design: exact match, then
# fuzzy match at 95% (flagged for confirmation), then **** for a
# person to assign. Raw AD/SQ Details names are never altered - only
# used as-is for lookup, per your instruction not to touch them.
def categorize(df, category_map_path):
    from rapidfuzz import process as fuzz_process

    print("\n" + "="*50)
    print("CATEGORIZER")
    print("="*50)

    cat_map = pd.read_excel(category_map_path)
    cat_map.columns = cat_map.columns.str.strip()
    cat_map["AD Name"]  = cat_map["AD Name"].astype(str).str.strip().str.upper()
    cat_map["Category"] = cat_map["Category"].astype(str).str.strip()

    lookup      = dict(zip(cat_map["AD Name"], cat_map["Category"]))
    lookup_keys = list(lookup.keys())

    exact_count, fuzzy_count, unknown_count = 0, 0, 0
    fuzzy_matches = []
    categories    = []

    for idx, row in df.iterrows():
        ad_name = str(row["AD/SQ Details"]).strip().upper()

        if ad_name in lookup:
            categories.append(lookup[ad_name])
            exact_count += 1
        else:
            result = fuzz_process.extractOne(ad_name, lookup_keys, score_cutoff=95)
            if result is not None:
                matched_name, score = result[0], result[1]
                suggested_category  = lookup[matched_name]
                categories.append(suggested_category)
                fuzzy_count += 1
                fuzzy_matches.append({
                    "Row": idx, "Original Name": ad_name,
                    "Matched To": matched_name,
                    "Suggested Category": suggested_category,
                    "Confidence": f"{round(score, 1)}%"
                })
                log_flag(
                    row_ref=f"Row {idx} ({row['Date']})", detail=ad_name,
                    issue=f"Fuzzy matched to '{matched_name}' ({round(score,1)}%) - assigned {suggested_category}",
                    suggestion="Confirm match is correct in Mapping Manager"
                )
            else:
                categories.append("****")
                unknown_count += 1
                log_flag(
                    row_ref=f"Row {idx} ({row['Date']})", detail=ad_name,
                    issue="No category match found",
                    suggestion="Assign category in Mapping Manager"
                )

    df["Aired Category"] = categories

    print(f"   Exact matches: {exact_count}  |  Fuzzy matches: {fuzzy_count}  |  Unknown (****): {unknown_count}")
    if unknown_count > 0:
        print(f"   [WARNING] {unknown_count} ads could not be categorized")
    print("="*50)

    return df, fuzzy_matches

# ── VALIDATOR ────────────────────────────────────────────────────
def validate(df):
    print("\n" + "="*50)
    print("VALIDATOR")
    print("="*50)

    issues_found = 0

    for idx, row in df.iterrows():
        ad_name  = str(row["AD/SQ Details"]).strip()
        ad_sq    = str(row["AD/SQ"]).strip()
        category = str(row["Aired Category"]).strip()
        seconds  = row["Seconds Aired"]
        time     = str(row["Time Aired"]).strip()
        date     = str(row["Date"]).strip()

        if ad_sq not in ["AD", "SQ"]:
            log_flag(f"Row {idx} ({date})", ad_name,
                     f"AD/SQ value is '{ad_sq}' - expected AD or SQ",
                     "Correct AD/SQ value manually")
            issues_found += 1

        expected_seconds = config.AD_SECONDS if ad_sq == "AD" else config.SQ_SECONDS
        if seconds != expected_seconds:
            original = seconds
            df.at[idx, "Seconds Aired"] = expected_seconds
            log_correction(f"Row {idx} ({date})", f"{original} seconds",
                           f"{expected_seconds} seconds",
                           f"Auto-corrected: {ad_sq} must be {expected_seconds} seconds")

        if category not in config.STANDARD_CATEGORIES and category != "****":
            log_flag(f"Row {idx} ({date})", ad_name,
                     f"Non-standard category: '{category}'",
                     "Reassign in Mapping Manager")
            issues_found += 1

        if not ad_name or ad_name.lower() in ["nan", "none", ""]:
            log_flag(f"Row {idx} ({date})", "BLANK",
                     "AD/SQ Details is blank",
                     "Review source file for this row")
            issues_found += 1

        if not time or time.lower() in ["nan", "none", ""]:
            log_flag(f"Row {idx} ({date})", ad_name,
                     "Time Aired is missing",
                     "This row should have been dropped by parser")
            issues_found += 1

        duplicates = df[
            (df["Date"]          == row["Date"]) &
            (df["AD/SQ Details"] == row["AD/SQ Details"]) &
            (df["Time Aired"]    == row["Time Aired"]) &
            (df["AD/SQ"]         == row["AD/SQ"]) &
            (df.index            != idx)
        ]
        if not duplicates.empty:
            log_flag(f"Row {idx} ({date})", ad_name,
                     f"Duplicate of row(s) {list(duplicates.index)}",
                     "Review and remove duplicate if confirmed")
            issues_found += 1

    print(f"   Rules checked: 6  |  Issues found: {issues_found}  |  "
          f"Corrections: {len(corrections_log)}  |  Flags: {len(flags_log)}")
    print("="*50)

    return df

def style_warehouse_workbook(wb):
    """Applies the standard warehouse look (header fill, alternating
    rows, column widths) to every sheet in the workbook. Shared by
    write_warehouse() AND any other code path that writes directly to
    warehouse.xlsx (e.g. app.py's Assign Categories / Mapping Manager
    saves) - so styling never quietly drifts out of sync again."""
    from openpyxl.styles import PatternFill, Font, Alignment

    header_fill = PatternFill(start_color="0090F0", end_color="0090F0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill    = PatternFill(start_color="E8F5FC", end_color="E8F5FC", fill_type="solid")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill
        for col in ws.columns:
            max_len, col_letter = 0, None
            for cell in col:
                if hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

# ── WAREHOUSE WRITER ─────────────────────────────────────────────
def write_warehouse(df, corrections, flags, dropped_rows, day_detection):
    print("\n" + "="*50)
    print("WAREHOUSE WRITER")
    print("="*50)

    warehouse_file = WAREHOUSE_PATH / "warehouse.xlsx"

    try:
        WAREHOUSE_PATH.mkdir(exist_ok=True)
    except OSError:
        print("[WARNING] Cannot create warehouse directory - likely running on Streamlit Cloud")
        print("[INFO] Persistence layer (GitHub commit-back) handles this - see app.py")
        return "memory_only"

    # Calendar-week label per row, NOT min/max of this upload alone -
    # this is the fix that makes daily uploads merge into one real
    # week instead of each day becoming its own tiny "week".
    df["Week"] = df["Full Date"].apply(
        lambda d: calendar_week_label(d) if d else "Unknown"
    )
    df["Processed On"] = datetime.now().strftime("%d.%m.%Y %H:%M")

    col_order = [
        "Week", "Date", "Full Date", "AD/SQ Details", "Time Aired",
        "Time Block", "Seconds Aired", "Aired Category", "AD/SQ",
        "Processed On"
    ]
    df = df[[c for c in col_order if c in df.columns]]

    # Day detection also gets a Week label, so it merges/replaces the
    # same way Clean Data does when a specific day is re-uploaded.
    detection_df = pd.DataFrame(day_detection) if day_detection else pd.DataFrame(
        columns=["Day", "Column (as in Excel)", "Row (as in Excel)",
                 "Date Assigned", "Full Date", "Raw Entries Scanned"])
    if not detection_df.empty:
        detection_df["Week"] = detection_df["Full Date"].apply(
            lambda d: calendar_week_label(d) if d else "Unknown"
        )

    # Same for dropped rows - without a Week/Day, a per-week export
    # couldn't scope the reconciliation to just that week, and
    # re-uploading a day would duplicate its dropped-row entries
    # instead of replacing them.
    dropped_df = pd.DataFrame(dropped_rows) if dropped_rows else pd.DataFrame(
        columns=["Row Reference", "AD/SQ Details", "AD/SQ", "Reason", "Day", "Full Date"])
    if not dropped_df.empty:
        dropped_df["Week"] = dropped_df["Full Date"].apply(
            lambda d: calendar_week_label(d) if d else "Unknown"
        )

    weeks_touched = df["Week"].unique()

    if warehouse_file.exists():
        print("   Existing warehouse found - merging...")
        existing_data       = pd.read_excel(warehouse_file, sheet_name="Clean Data")
        existing_corrections = pd.read_excel(warehouse_file, sheet_name="Corrections Log")
        existing_flags       = pd.read_excel(warehouse_file, sheet_name="Flags Log")
        try:
            existing_dropped = pd.read_excel(warehouse_file, sheet_name="Dropped Rows Log")
        except Exception:
            existing_dropped = pd.DataFrame(
                columns=["Row Reference", "AD/SQ Details", "AD/SQ", "Reason", "Day", "Full Date", "Week"])
        try:
            existing_detection = pd.read_excel(warehouse_file, sheet_name="Day Detection Log")
        except Exception:
            existing_detection = pd.DataFrame(
                columns=["Day", "Column (as in Excel)", "Row (as in Excel)",
                         "Date Assigned", "Full Date", "Raw Entries Scanned", "Week"])

        # Replace matching (Week, Date) pairs - this preserves other
        # days already in the same week when a single day is
        # re-uploaded, instead of wiping the whole week.
        if "Week" in existing_data.columns and "Date" in existing_data.columns:
            mask = existing_data.apply(
                lambda r: (r["Week"], r["Date"]) in set(zip(df["Week"], df["Date"])),
                axis=1
            )
            replaced = int(mask.sum())
            if replaced:
                print(f"   [INFO] Replacing {replaced} existing row(s) for the day(s) being re-uploaded")
            existing_data = existing_data[~mask]

        if not existing_detection.empty and not detection_df.empty:
            det_mask = existing_detection.apply(
                lambda r: (r["Week"], r["Day"]) in set(zip(detection_df["Week"], detection_df["Day"])),
                axis=1
            )
            existing_detection = existing_detection[~det_mask]

        if not existing_dropped.empty and not dropped_df.empty and "Week" in existing_dropped.columns:
            drop_mask = existing_dropped.apply(
                lambda r: (r["Week"], r["Day"]) in set(zip(dropped_df["Week"], dropped_df["Day"])),
                axis=1
            )
            existing_dropped = existing_dropped[~drop_mask]

        combined_data = pd.concat([existing_data, df], ignore_index=True)
        combined_corrections = pd.concat(
            [existing_corrections, pd.DataFrame(corrections)], ignore_index=True
        ) if corrections else existing_corrections
        combined_flags = pd.concat(
            [existing_flags, pd.DataFrame(flags)], ignore_index=True
        ) if flags else existing_flags
        combined_dropped = pd.concat(
            [existing_dropped, dropped_df], ignore_index=True
        ) if dropped_rows else existing_dropped
        combined_detection = pd.concat([existing_detection, detection_df], ignore_index=True)
    else:
        print("   No existing warehouse - creating new...")
        combined_data = df
        combined_corrections = pd.DataFrame(corrections) if corrections else pd.DataFrame(
            columns=["Row Reference", "Original Value", "Corrected To", "Reason"])
        combined_flags = pd.DataFrame(flags) if flags else pd.DataFrame(
            columns=["Row Reference", "AD/SQ Details", "Issue", "Suggestion"])
        combined_dropped = dropped_df
        combined_detection = detection_df

    from openpyxl.worksheet.datavalidation import DataValidation

    with pd.ExcelWriter(warehouse_file, engine="openpyxl") as writer:
        combined_data.to_excel(writer, index=False, sheet_name="Clean Data")
        combined_corrections.to_excel(writer, index=False, sheet_name="Corrections Log")
        combined_flags.to_excel(writer, index=False, sheet_name="Flags Log")
        combined_dropped.to_excel(writer, index=False, sheet_name="Dropped Rows Log")
        combined_detection.to_excel(writer, index=False, sheet_name="Day Detection Log")

        weeks_in_warehouse = combined_data["Week"].unique() if "Week" in combined_data.columns else []
        summary_data = {
            "Metric": ["Total Rows in Warehouse", "Total Weeks Stored", "Total ADs",
                       "Total SQs", "Total Corrections", "Total Flags",
                       "Total Dropped Rows", "Last Updated"],
            "Value": [len(combined_data), len(weeks_in_warehouse),
                      len(combined_data[combined_data["AD/SQ"] == "AD"]),
                      len(combined_data[combined_data["AD/SQ"] == "SQ"]),
                      len(combined_corrections), len(combined_flags),
                      len(combined_dropped),
                      datetime.now().strftime("%d.%m.%Y %H:%M")]
        }
        pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name="Summary")

        wb = writer.book
        style_warehouse_workbook(wb)

        ws_clean = wb["Clean Data"]
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(config.STANDARD_CATEGORIES) + '"',
            allow_blank=True, showDropDown=False
        )
        dv.sqref = f"H2:H{len(combined_data) + 1}"
        ws_clean.add_data_validation(dv)

    print(f"   Week(s) touched: {list(weeks_touched)}")
    print(f"   Rows added: {len(df)}  |  Total in warehouse: {len(combined_data)}")
    print(f"   Dropped rows logged this run: {len(dropped_rows)}")
    print(f"   Saved to: {warehouse_file}")
    print("="*50)

    return True

# ── MAIN ─────────────────────────────────────────────────────────
if __name__ == "__main__":

    raw_files = list(RAW_FOLDER.glob("*.xlsx"))
    if not raw_files:
        print("[ERROR] No .xlsx file found in raw_inputs folder.")
        exit()

    raw_file = max(raw_files, key=lambda f: f.stat().st_mtime)
    print(f"\nProcessing: {raw_file.name}")

    clear_logs()
    result = preflight_check(raw_file)
    if not result:
        print("\n[ERROR] Pre-flight failed.")
        exit()

    df_raw, days_found = result
    df_parsed = parse_raw_file(df_raw, days_found)

    if df_parsed is not None:
        df_categorized, fuzzy_matches = categorize(df_parsed, CATEGORY_MAP)
        df_validated = validate(df_categorized)

        print("\nFinal category breakdown:")
        print(df_validated["Aired Category"].value_counts().to_string())
        print(f"\nCorrections log: {len(corrections_log)} entries")
        print(f"Flags log: {len(flags_log)} entries")
        print(f"Dropped rows log: {len(dropped_rows_log)} entries")

        success = write_warehouse(df_validated.copy(), corrections_log, flags_log,
                                   dropped_rows_log, day_detection_log)

        if success:
            print("\n[OK] Pipeline complete. Warehouse updated successfully.")
        else:
            print("\n[WARNING] Warehouse write skipped.")
